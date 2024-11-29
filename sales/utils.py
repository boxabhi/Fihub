from openpyxl.utils import get_column_letter
import openpyxl
from django.conf import settings
import pandas as pd
import os,sys
import numpy as np
from openpyxl.utils import get_column_letter




class PlainExcelGenerator:
    """"
        A utility function for generating Excel sheet
    """

    css_alt_rows = 'background-color: white; color: black;width:500px;'
    css_indexes = 'background-color: #0354a3; color: white;width:500px;'
    all_css = 'font-size:23px;width:400px;margin:40px'
    file_path = 'public/static/excels'

    def __init__(self, datas, file_name) -> None:
        self.datas = datas
        self.file_name = file_name


    def set_column_widths(self, df, worksheet):
        for i, col in enumerate(df.columns, 1):  # 1-based index
            max_length = max(df[col].astype(str).map(len).max(), len(str(col))) + 2  # Adding 2 for padding
            worksheet.column_dimensions[get_column_letter(i)].width = max_length

    def generateExcel(self):
        try:
            df = pd.DataFrame(self.datas)
            styled_df = (df.style
                            .apply(lambda col: np.where(col.index % 2, self.css_alt_rows, None))  # alternating rows
                            .applymap_index(lambda _: self.css_indexes, axis=0)  # row indexes (pandas 1.4.0+)
                            .applymap_index(lambda _: self.css_indexes, axis=1)  # col indexes (pandas 1.4.0+)
                            .applymap_index(lambda _: self.all_css)
                        )
            
            file_path = f'{settings.BASE_DIR}/{self.file_path}/{self.file_name}.xlsx'
            styled_df.to_excel(file_path, engine='openpyxl', index=False)
            
            # Load the workbook to adjust column widths
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active

            # Set the column widths
            self.set_column_widths(df, worksheet)

            # Save the updated workbook
            workbook.save(file_path)
            workbook.close()

            return True, f"/media/excels/{self.file_name}.xlsx"
        except Exception as e:
            print(e)
            exc_type, exc_obj, exc_tb = sys.exc_info()
            fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
            print(exc_type, fname, exc_tb.tb_lineno)
            print(e)
            return False, f"{exc_type, fname, exc_tb.tb_lineno,exit}"
        